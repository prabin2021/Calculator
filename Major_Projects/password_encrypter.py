import os
import hashlib   # this library is used to encrypt the password in hashkey
from tkinter  import *     # for gui (Graphical User Interface)
from tkinter import messagebox
from openpyxl import Workbook , load_workbook      # for manipulating the excel file

file_path = "C:/Users/sigde/OneDrive/Desktop/Python/Major_Projects/Users.xlsx"

def encrypt_password(password):
    encrypted_password =  hashlib.sha256(password.encode()).hexdigest()
    return encrypted_password

def create_file():
    if not os.path.exists(file_path):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Users"
        sheet.append(["Username","Password"])
        wb.save(file_path)

def register_user():
    
    username = entry_username.get()
    password = entry_password.get()
    
    if username == "" or password == "":
        print("Empty username or password")
        messagebox.showerror("Error","All fields are required")
        
    encrypted_password = encrypt_password(password)
    wb = load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows(min_row = 2, values_only= True):
        if row[0] == username:
            messagebox.showerror("Error","User name already exists")
            clear_fields()
            return
    sheet.append([username,encrypted_password])
    wb.save(file_path)
    messagebox.showinfo("Success","User registered sucessfully")
    clear_fields()
    
def login_user():
    
    username = entry_username.get()
    password = entry_password.get()
    
    encrypted_password = encrypt_password(password)
        
    wb = load_workbook(file_path)
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=2, values_only= True):
        if row[0] == username and row[1] ==  encrypted_password:
            messagebox.showinfo("Sucess","Login succesful")
            clear_fields()
            return
    messagebox.showerror("Error","Login failed") 
    clear_fields() 
            
def clear_fields(): 
    entry_username.delete(0,END)
    entry_password.delete(0,END)
    

root = Tk()
root.title("Secured Login System")
root.geometry("650x450")
root.resizable(False,False)


Label(root,width=30 ,text = "Username").pack(pady=10)
entry_username = Entry(root)
entry_username.pack()

Label(root,width=30, text = "Password").pack(pady=10)
entry_password = Entry(root, show = "*",)
entry_password.pack()
        
Button(root, text = "Register", width= 20 ,command= register_user).pack(pady=30)
        
Button(root, text = "Login", width= 20 ,command= login_user).pack(pady=20)


root.mainloop()